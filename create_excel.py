"""
建立信用卡資料 Excel 檔案
從 CSV 轉換為 Excel,解決 Excel 打開 CSV 格式錯亂的問題
"""
import pandas as pd
from pathlib import Path

# 讀取 CSV
csv_path = Path(__file__).parent / "信用卡資料模板.csv"
excel_path = Path(__file__).parent / "信用卡資料模板.xlsx"

print(f"📖 讀取 CSV: {csv_path}")
df = pd.read_csv(csv_path, encoding='utf-8-sig')

print(f"✅ 載入 {len(df)} 張信用卡資料")

# 建立 Excel
print(f"\n📝 建立 Excel: {excel_path}")

with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='信用卡資料')
    
    worksheet = writer.sheets['信用卡資料']
    
    # 調整欄寬
    worksheet.column_dimensions['A'].width = 25  # 信用卡名稱
    worksheet.column_dimensions['B'].width = 15  # 銀行
    worksheet.column_dimensions['C'].width = 80  # 回饋方案
    worksheet.column_dimensions['D'].width = 40  # APP切換方案
    worksheet.column_dimensions['E'].width = 15  # 回饋開始日
    worksheet.column_dimensions['F'].width = 15  # 回饋到期日
    worksheet.column_dimensions['G'].width = 10  # 年費
    worksheet.column_dimensions['H'].width = 40  # 備註
    
    # 設定標題列樣式
    from openpyxl.styles import Font, PatternFill, Alignment
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 設定內容自動換行
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')

print(f"✅ Excel 檔案建立完成!")
print(f"\n📊 檔案位置: {excel_path.absolute()}")
print(f"\n💡 現在可以用 Excel 正常打開 {excel_path.name} 了!")
